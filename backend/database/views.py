from django.db.models import Sum, Count
from django.utils import timezone
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import viewsets, status
from django_filters.rest_framework import DjangoFilterBackend
from . import models,serializer
from django.http import HttpResponse
from decimal import Decimal    
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class personaView(viewsets.ModelViewSet):
    queryset = models.Persona.objects.all()
    serializer_class = serializer.personaSerializer

class salidaView(viewsets.ModelViewSet):
    queryset = models.Salida.objects.all().order_by('-fecha')

    def get_serializer_class(self):
        if self.action in ['create','update','partial_update']:
            return serializer.SalidaWriteSerializer
        return serializer.SalidaReadSerializer

class deliveristaView(viewsets.ModelViewSet):
    queryset = models.Trabajador.objects.all()
    serializer_class = serializer.DeliveristaSerializer

class productoView(viewsets.ModelViewSet):
    queryset = models.Producto.objects.all()
    serializer_class = serializer.ProductoSerializer
    
class retornoView(viewsets.ModelViewSet):
    queryset = models.Retorno.objects.all()
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id_salida'] 
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return serializer.RetornoWriteSerializer
        return serializer.RetornoReadSerializer

    def get_object(self):
        if self.action in ['retrieve', 'update', 'partial_update', 'destroy']:
            try:
                id_salida = self.kwargs['pk']
                return models.Retorno.objects.get(id_salida=id_salida)
            except models.Retorno.DoesNotExist:
                from rest_framework.exceptions import NotFound
                raise NotFound("No se encontró retorno para la salida con ID {}".format(id_salida))
            except ValueError:
                from rest_framework.exceptions import ValidationError
                raise ValidationError("ID de salida inválido")

        return super().get_object()
    
class RetornoViewRes(viewsets.ModelViewSet):
    queryset = models.Retorno.objects.all()
    serializer_class = serializer.RetornoReadSerializer

    @action(detail=False, methods=['get'])
    def resumen_caja_completo(self, request):
        hoy = timezone.now().date()
        print(f"Calculando resumen para: {hoy}")
        
        # 1. INGRESOS POR RETORNOS (Dinero que YA pagaron)
        retornos_hoy = models.Retorno.objects.filter(id_salida__fecha=hoy)
        
        total_ingresos_reales = retornos_hoy.aggregate(
            total=Sum('total_cancelado')
        )['total'] or 0
        
        # 2. SALIDAS POR TIPO DE TRABAJADOR (Lo que se llevaron para vender)
        salidas_hoy = models.Salida.objects.filter(fecha=hoy)
        
        # Delivery - Solo repartidores (ID 3)
        salidas_delivery = salidas_hoy.filter(id_trabajador__id_tipo_trabajador=3)
        total_delivery = salidas_delivery.aggregate(
            total=Sum('total_cancelar')
        )['total'] or 0
        
        # Venta local - Jefe de planta (1) y Asistente (2)
        salidas_local = salidas_hoy.filter(id_trabajador__id_tipo_trabajador__in=[1, 2])
        total_local = salidas_local.aggregate(
            total=Sum('total_cancelar')
        )['total'] or 0
        
        total_egresos = models.CajaIe.objects.filter(
            tipo="egreso",
            fecha__date=hoy  # ← CAMBIO AQUÍ: usa __date para extraer solo la fecha
        ).aggregate(
            total=Sum('nonto')
        )['total'] or 0

        # Debug info
        print(f"Salidas delivery hoy: {salidas_delivery.count()}")
        print(f"Salidas local hoy: {salidas_local.count()}")
        print(f"Retornos hoy: {retornos_hoy.count()}")
        print(f"Egresos hoy: {total_egresos}")

        data = {
            'total_hoy': float(total_ingresos_reales),
            'total_repartidores': float(total_delivery),
            'total_no_repartidores': float(total_local),
            'total_egresos': float(total_egresos),
            'balance_neto': float(total_ingresos_reales - total_egresos)
        }
        
        return Response(data)

class TrabajadorViewSet(viewsets.ModelViewSet):
    queryset = models.Trabajador.objects.select_related(
        'id_persona', 
        'id_tipo_trabajador', 
        'id_horario'
    ).all()
    
    def get_serializer_class(self):
        """Retorna el serializer apropiado según la acción"""
        if self.action == 'list':
            return serializer.TrabajadorReadSerializer
        elif self.action == 'retrieve':
            return serializer.TrabajadorReadSerializer
        elif self.action == 'create':
            return serializer.TrabajadorCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return serializer.TrabajadorUpdateSerializer
        return serializer.TrabajadorReadSerializer
    
    def create(self, request, *args, **kwargs):
        """Crear trabajador con persona"""
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        trabajador = ser.save()

        read_serializer = serializer.TrabajadorReadSerializer(trabajador)
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        """Actualizar trabajador y persona"""
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        ser = self.get_serializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        trabajador = ser.save()

        read_serializer = serializer.TrabajadorReadSerializer(trabajador)
        return Response(read_serializer.data)
    
    def destroy(self, request, *args, **kwargs):
        """Eliminar trabajador con manejo de errores mejorado"""
        try:
            instance = self.get_object()
            persona = instance.id_persona
            
            # Verificar si el trabajador tiene operaciones relacionadas
            tiene_ventas = models.Venta.objects.filter(id_trabajador=instance).exists()
            tiene_salidas = models.Salida.objects.filter(id_trabajador=instance).exists()
            
            if tiene_ventas or tiene_salidas:
                return Response(
                    {"error": "No se puede eliminar el trabajador porque tiene operaciones registradas. Considere desactivarlo en lugar de eliminarlo."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Eliminar trabajador y persona
            self.perform_destroy(instance)
            persona.delete()  # Eliminar también la persona asociada
            
            return Response(
                {"message": "Trabajador eliminado exitosamente"},
                status=status.HTTP_200_OK
            )
        
        except Exception as e:
            print(f"Error eliminando trabajador: {str(e)}")
            return Response(
                {"error": f"Error al eliminar trabajador: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def por_tipo(self, request):
        """Filtrar trabajadores por tipo"""
        tipo_id = request.query_params.get('tipo_id', None)
        if tipo_id:
            trabajadores = self.queryset.filter(id_tipo_trabajador=tipo_id)
            ser = serializer.TrabajadorListSerializer(trabajadores, many=True)
            return Response(ser.data)
        return Response(
            {"error": "Debe proporcionar tipo_id como parámetro"}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    @action(detail=True, methods=['get'])
    def detalle_completo(self, request, pk=None):
        """Obtener detalle completo del trabajador"""
        trabajador = self.get_object()
        ser = serializer.TrabajadorReadSerializer(trabajador)
        return Response(ser.data)
    
    @action(detail=False, methods=['post'])
    def generar_vista_previa(self, request):
        """Endpoint para vista previa de trabajadores"""
        try:
            from . import serializer as app_serializer
            reporte_serializer = app_serializer.ReporteFiltrosSerializer(data=request.data)
            
            if not reporte_serializer.is_valid():
                return Response(reporte_serializer.errors, status=400)
            
            data = reporte_serializer.validated_data
            return self._generar_vista_previa_trabajadores(data)
            
        except Exception as e:
            print(f"❌ ERROR en vista previa trabajadores: {str(e)}")
            return Response({"error": f"Error interno: {str(e)}"}, status=500)

    def _generar_vista_previa_trabajadores(self, filtros):
        """Vista previa para reporte de trabajadores"""
        trabajadores = models.Trabajador.objects.all().select_related(
            'id_persona', 'id_tipo_trabajador'
        )
        
        if filtros.get('id_trabajador'):
            trabajadores = trabajadores.filter(id_trabajador=filtros['id_trabajador'])
        
        datos_tabla = []
        grafico_trabajadores = []
        
        for trabajador in trabajadores[:10]:  # Solo primeros 10 para vista previa
            ventas_trabajador = models.Venta.objects.filter(id_trabajador=trabajador)
            salidas_trabajador = models.Salida.objects.filter(id_trabajador=trabajador)
            
            # Aplicar filtros de fecha si existen
            if filtros.get('fecha_inicio'):
                ventas_trabajador = ventas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
                salidas_trabajador = salidas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
            if filtros.get('fecha_fin'):
                ventas_trabajador = ventas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
                salidas_trabajador = salidas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
            
            total_ventas = ventas_trabajador.aggregate(Sum('total_cancelado'))['total_cancelado__sum'] or 0
            total_entregas = salidas_trabajador.count()
            entregas_completadas = salidas_trabajador.filter(id_estado_salida=2).count()
            
            eficiencia = 0
            if total_entregas > 0:
                eficiencia = (entregas_completadas / total_entregas) * 100
            
            datos_tabla.append({
                'nombre': f"{trabajador.id_persona.nombre_p} {trabajador.id_persona.apellido_p}",
                'dni': trabajador.id_persona.dni_p,
                'tipo_trabajador': trabajador.id_tipo_trabajador.nom_tt,
                'total_ventas': float(total_ventas),
                'total_entregas': total_entregas,
                'entregas_completadas': entregas_completadas,
                'eficiencia': round(eficiencia, 2),
            })
            
            grafico_trabajadores.append({
                'nombre': f"{trabajador.id_persona.nombre_p} {trabajador.id_persona.apellido_p}",
                'eficiencia': round(eficiencia, 2)
            })
        
        return Response({
            'total_registros': trabajadores.count(),
            'promedio_eficiencia': round(sum([t['eficiencia'] for t in grafico_trabajadores]) / len(grafico_trabajadores) if grafico_trabajadores else 0, 2),
            'headers': ['NOMBRE', 'DNI', 'TIPO TRABAJADOR', 'TOTAL VENTAS', 'TOTAL ENTREGAS', 'ENTREGAS COMPLETADAS', 'EFICIENCIA (%)'],
            'datos': datos_tabla,
            'grafico_trabajadores': grafico_trabajadores
        })
    
class TipoTrabajadorViewSet(viewsets.ModelViewSet):
    queryset = models.TipoTrabajador.objects.all()
    serializer_class = serializer.TipoTrabajadorSerializer

class HorarioViewSet(viewsets.ModelViewSet):
    queryset = models.Horario.objects.all()
    serializer_class = serializer.HorarioSerializer

class DashboardViewSet(viewsets.ViewSet):
    
    def list(self, request):
        return self.estadisticas_principales(request)
    
    @action(detail=False, methods=['get'])
    def estadisticas_principales(self, request):
        hoy = timezone.now().date()
        ayer = hoy - timezone.timedelta(days=1)

        entregas_hoy = models.Salida.objects.filter(fecha=hoy).count()
        entregas_ayer = models.Salida.objects.filter(fecha=ayer).count()
        crecimiento_entregas = self._calcular_crecimiento(entregas_hoy, entregas_ayer)

        ingresos_hoy = models.Venta.objects.filter(fecha=hoy).aggregate(
            total=Sum('total_cancelado')
        )['total'] or 0
        
        ingresos_ayer = models.Venta.objects.filter(fecha=ayer).aggregate(
            total=Sum('total_cancelado')
        )['total'] or 0
        crecimiento_ingresos = self._calcular_crecimiento(ingresos_hoy, ingresos_ayer)

        total_trabajadores = models.Trabajador.objects.count()
        trabajadores_activos = models.User.objects.filter(estado=True).count()
        estado_trabajadores = "Todos activos" if trabajadores_activos == total_trabajadores else f"{trabajadores_activos} activos"
        
        data = {
            'entregas_hoy': entregas_hoy,
            'crecimiento_entregas': crecimiento_entregas,
            'ingresos_hoy': ingresos_hoy,
            'crecimiento_ingresos': crecimiento_ingresos,
            'total_trabajadores': total_trabajadores,
            'trabajadores_activos': trabajadores_activos,
            'estado_trabajadores': estado_trabajadores
        }

        dashboard_serializer = serializer.DashboardSerializer(data)
        return Response(dashboard_serializer.data)
    
    def _calcular_crecimiento(self, valor_hoy, valor_ayer):
        if valor_ayer == 0:
            return "+100%" if valor_hoy > 0 else "0%"
        
        crecimiento = ((valor_hoy - valor_ayer) / valor_ayer) * 100
        signo = "+" if crecimiento > 0 else ""
        return f"{signo}{crecimiento:.0f}%"
    
    @action(detail=False, methods=['get'])
    def performance_entregas(self, request):
        from datetime import timedelta
        
        hoy = timezone.now().date()
        ultimos_7_dias = [hoy - timedelta(days=i) for i in range(6, -1, -1)]
        
        datos_performance = []
        
        for fecha in ultimos_7_dias:
            entregas_programadas = models.Salida.objects.filter(fecha=fecha).count()

            entregas_completadas = models.Salida.objects.filter(
                fecha=fecha,
                id_estado_salida=2  
            ).count()

            retornos = models.Retorno.objects.filter(
                id_salida__fecha=fecha
            ).aggregate(
                total_retornos=Sum('cantidad')
            )['total_retornos'] or 0
 
            monto_pendiente = models.Salida.objects.filter(
                fecha=fecha,
                id_estado_pago__in=[2, 3]
            ).aggregate(
                total=Sum('total_cancelar')
            )['total'] or 0

            eficiencia = 0
            if entregas_programadas > 0:
                eficiencia = (entregas_completadas / entregas_programadas) * 100
            
            datos_performance.append({
                'fecha': fecha,
                'entregas_programadas': entregas_programadas,
                'entregas_completadas': entregas_completadas,
                'retornos': retornos,
                'eficiencia': round(eficiencia, 1),
                'monto_pendiente': float(monto_pendiente)
            })

        performance_serializer = serializer.PerformanceEntregasSerializer(datos_performance, many=True)
        return Response(performance_serializer.data)

class MovimientoCajaViewSet(viewsets.ViewSet):
    
    @action(detail=False, methods=['post'])
    def registrar_movimiento(self, request):
        try:
            from . import serializer as app_serializer
            
            movimiento_serializer = app_serializer.MovimientoCajaSerializer(data=request.data)
            if not movimiento_serializer.is_valid():
                return Response(movimiento_serializer.errors, status=400)
            
            data = movimiento_serializer.validated_data
            
            # Obtener el trabajador seleccionado (NO más el default)
            try:
                trabajador = models.Trabajador.objects.get(id_trabajador=data['id_trabajador'])
            except models.Trabajador.DoesNotExist:
                return Response({"error": "Trabajador no encontrado"}, status=400)
            
            movimiento = models.CajaIe.objects.create(
                tipo=data['tipo'],
                nonto=Decimal(str(data['monto'])),
                descripcion=data['descripcion'],
                id_trabajador=trabajador  # ← Usar el trabajador seleccionado
            )
            
            return Response({
                "message": "Movimiento registrado exitosamente",
                "id": movimiento.id_caja_ie,
                "tipo": movimiento.tipo,
                "monto": float(movimiento.nonto),
                "descripcion": movimiento.descripcion
            }, status=201)
            
        except Exception as e:
            print(f"Error registrando movimiento: {str(e)}")
            return Response({"error": str(e)}, status=500)

    # AÑADIR ESTE NUEVO ENDPOINT para obtener trabajadores
    @action(detail=False, methods=['get'])
    def obtener_trabajadores(self, request):
        try:
            trabajadores = models.Trabajador.objects.all().select_related('id_persona')
            
            datos = []
            for trab in trabajadores:
                datos.append({
                    'id_trabajador': trab.id_trabajador,
                    'nombre_completo': f"{trab.id_persona.nombre_p} {trab.id_persona.apellido_p}",
                    'dni': trab.id_persona.dni_p
                })
            
            return Response(datos)
            
        except Exception as e:
            return Response({"error": str(e)}, status=500)
    
    @action(detail=False, methods=['get'])
    def obtener_movimientos(self, request):
        try:
            movimientos = models.CajaIe.objects.all().select_related('id_trabajador').order_by('-id_caja_ie')[:50]
            
            datos = []
            for mov in movimientos:
                datos.append({
                    'id': mov.id_caja_ie,
                    'fecha': timezone.now().date(),
                    'tipo': mov.tipo,
                    'monto': float(mov.nonto),
                    'descripcion': mov.descripcion,
                    'metodo': 'efectivo',
                    'responsable': mov.id_trabajador.id_persona.nombre_p if mov.id_trabajador else 'N/A'
                })
            
            return Response(datos)
            
        except Exception as e:
            return Response({"error": str(e)}, status=500)
    
    @action(detail=False, methods=['get'])
    def movimientos_caja_completos(self, request):
        try:
   
            fecha_inicio = request.GET.get('fechaInicio')
            fecha_fin = request.GET.get('fechaFin')
            tipo = request.GET.get('tipo', 'todos')
            metodo = request.GET.get('metodo', 'todos')
            
            movimientos = []

            retornos = models.Retorno.objects.all()

            if fecha_inicio:
                retornos = retornos.filter(id_salida__fecha__gte=fecha_inicio)
            if fecha_fin:
                retornos = retornos.filter(id_salida__fecha__lte=fecha_fin)
                
            for retorno in retornos:
                movimientos.append({
                    'id': f"R_{retorno.id_retorno}",
                    'fecha': retorno.id_salida.fecha,
                    'hora': retorno.id_salida.hora,
                    'monto': float(retorno.total_cancelado),
                    'tipo': 'ingreso',
                    'metodo': self._determinar_metodo_pago(retorno.id_pago),
                    'descripcion': f"Venta - {retorno.id_salida.id_producto.nom_producto}",
                    'responsable': retorno.id_salida.id_trabajador.id_persona.nombre_p,
                    'origen': 'venta'
                })

            egresos = models.CajaIe.objects.all()

            if tipo == 'ingreso':
                egresos = egresos.none() 
            
            for egreso in egresos:
                movimientos.append({
                    'id': f"E_{egreso.id_caja_ie}",
                    'fecha': timezone.now().date(),
                    'hora': timezone.now().time(),
                    'monto': float(egreso.nonto),
                    'tipo': egreso.tipo, 
                    'metodo': 'efectivo',
                    'descripcion': egreso.descripcion or "Movimiento de caja",
                    'responsable': egreso.id_trabajador.id_persona.nombre_p if egreso.id_trabajador else 'Sistema',
                    'origen': 'caja_manual'
                })
            
            movimientos.sort(key=lambda x: (x['fecha'], x['hora']), reverse=True)

            if tipo in ['ingreso', 'egreso']:
                movimientos = [m for m in movimientos if m['tipo'] == tipo]

            if metodo != 'todos':
                movimientos = [m for m in movimientos if m['metodo'] == metodo]
            
            return Response(movimientos[:100])
            
        except Exception as e:
            print(f"Error obteniendo movimientos: {str(e)}")
            return Response({"error": str(e)}, status=500)

    def _determinar_metodo_pago(self, pago):
        """Determina el método de pago basado en los montos"""
        if pago.yape > 0 and pago.efectivo > 0:
            return 'mixto'
        elif pago.yape > 0:
            return 'yape'
        else:
            return 'efectivo'
        
    @action(detail=False, methods=['get'])
    def generar_reporte_excel(self, request):
        try:
            # Obtener parámetros de filtro
            fecha_inicio = request.GET.get('fechaInicio')
            fecha_fin = request.GET.get('fechaFin')
            tipo = request.GET.get('tipo', 'todos')
            metodo = request.GET.get('metodo', 'todos')
            
            # Aplicar los mismos filtros que en movimientos_caja_completos
            movimientos = []
            
            # Obtener retornos (ingresos) con filtros
            retornos = models.Retorno.objects.all()
            if fecha_inicio:
                retornos = retornos.filter(id_salida__fecha__gte=fecha_inicio)
            if fecha_fin:
                retornos = retornos.filter(id_salida__fecha__lte=fecha_fin)
                
            for retorno in retornos:
                movimientos.append({
                    'id': f"R_{retorno.id_retorno}",
                    'fecha': retorno.id_salida.fecha,
                    'hora': retorno.id_salida.hora,
                    'monto': float(retorno.total_cancelado),
                    'tipo': 'ingreso',
                    'metodo': self._determinar_metodo_pago(retorno.id_pago),
                    'descripcion': f"Venta - {retorno.id_salida.id_producto.nom_producto}",
                    'responsable': f"{retorno.id_salida.id_trabajador.id_persona.nombre_p} {retorno.id_salida.id_trabajador.id_persona.apellido_p}",
                    'origen': 'venta'
                })

            # Obtener movimientos manuales de caja con filtros
            egresos = models.CajaIe.objects.all()
            if fecha_inicio:
                egresos = egresos.filter(fecha__date__gte=fecha_inicio)
            if fecha_fin:
                egresos = egresos.filter(fecha__date__lte=fecha_fin)
            
            if tipo == 'ingreso':
                egresos = egresos.none()
            elif tipo == 'egreso':
                egresos = egresos.filter(tipo='egreso')
            
            for movimiento in egresos:
                movimientos.append({
                    'id': f"E_{movimiento.id_caja_ie}",
                    'fecha': movimiento.fecha.date() if hasattr(movimiento.fecha, 'date') else movimiento.fecha,
                    'hora': movimiento.fecha.time() if hasattr(movimiento.fecha, 'time') else movimiento.fecha,
                    'monto': float(movimiento.nonto),
                    'tipo': movimiento.tipo,
                    'metodo': 'efectivo',  # O puedes agregar campo método en CajaIe
                    'descripcion': movimiento.descripcion or "Movimiento de caja",
                    'responsable': f"{movimiento.id_trabajador.id_persona.nombre_p} {movimiento.id_trabajador.id_persona.apellido_p}" if movimiento.id_trabajador else 'Sistema',
                    'origen': 'caja_manual'
                })

            # Aplicar filtros adicionales
            if tipo in ['ingreso', 'egreso']:
                movimientos = [m for m in movimientos if m['tipo'] == tipo]
            
            if metodo != 'todos':
                movimientos = [m for m in movimientos if m['metodo'] == metodo]
            
            # Ordenar por fecha y hora
            movimientos.sort(key=lambda x: (x['fecha'], x['hora']), reverse=True)
            
            # Crear el archivo Excel profesional
            return self._crear_excel_profesional(movimientos, fecha_inicio, fecha_fin, tipo, metodo)
            
        except Exception as e:
            print(f"Error generando reporte: {str(e)}")
            return Response({"error": str(e)}, status=500)

    def _crear_excel_profesional(self, movimientos, fecha_inicio, fecha_fin, tipo, metodo):
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Caja"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
        title_font = Font(bold=True, size=16, color="2C5F2D")
        subheader_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título principal
        ws.merge_cells('A1:H1')
        ws['A1'] = "REPORTE DE CAJA - YACU SELVA"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Información del reporte
        ws.merge_cells('A2:H2')
        periodo = "PERIODO: "
        if fecha_inicio and fecha_fin:
            periodo += f"{fecha_inicio} al {fecha_fin}"
        elif fecha_inicio:
            periodo += f"Desde {fecha_inicio}"
        elif fecha_fin:
            periodo += f"Hasta {fecha_fin}"
        else:
            periodo += "Todos los registros"
            
        if tipo != 'todos':
            periodo += f" | TIPO: {tipo.upper()}"
        if metodo != 'todos':
            periodo += f" | MÉTODO: {metodo.upper()}"
            
        ws['A2'] = periodo
        ws['A2'].font = subheader_font
        ws['A2'].alignment = center_align
        
        # Fecha de generación
        ws.merge_cells('A3:H3')
        ws['A3'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].alignment = center_align
        
        # Espacio
        ws.row_dimensions[4].height = 5
        
        # Encabezados de columnas
        headers = ['FECHA', 'HORA', 'TIPO', 'DESCRIPCIÓN', 'RESPONSABLE', 'MÉTODO', 'MONTO (S/.)', 'ORIGEN']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Datos
        row_num = 6
        total_ingresos = 0
        total_egresos = 0
        
        for mov in movimientos:
            # Formatear fecha y hora
            fecha_str = mov['fecha'].strftime('%Y-%m-%d') if hasattr(mov['fecha'], 'strftime') else str(mov['fecha'])
            hora_str = mov['hora'].strftime('%H:%M:%S') if hasattr(mov['hora'], 'strftime') else str(mov['hora'])
            
            # Determinar color por tipo
            tipo_color = "00B050" if mov['tipo'] == 'ingreso' else "FF0000"
            tipo_texto = "INGRESO" if mov['tipo'] == 'ingreso' else "EGRESO"
            
            # Escribir datos
            ws.cell(row=row_num, column=1, value=fecha_str).border = border
            ws.cell(row=row_num, column=2, value=hora_str).border = border
            ws.cell(row=row_num, column=3, value=tipo_texto).border = border
            ws.cell(row=row_num, column=4, value=mov['descripcion']).border = border
            ws.cell(row=row_num, column=5, value=mov['responsable']).border = border
            ws.cell(row=row_num, column=6, value=mov['metodo'].upper()).border = border
            ws.cell(row=row_num, column=7, value=mov['monto']).border = border
            ws.cell(row=row_num, column=8, value=mov['origen']).border = border
            
            # Aplicar color al tipo
            tipo_cell = ws.cell(row=row_num, column=3)
            tipo_cell.font = Font(bold=True, color=tipo_color)
            
            # Aplicar color al monto
            monto_cell = ws.cell(row=row_num, column=7)
            monto_cell.font = Font(bold=True, color=tipo_color)
            monto_cell.number_format = '#,##0.00'
            
            # Sumar totales
            if mov['tipo'] == 'ingreso':
                total_ingresos += mov['monto']
            else:
                total_egresos += mov['monto']
            
            row_num += 1
        
        # Totales
        ws.merge_cells(f'A{row_num+1}:F{row_num+1}')
        total_cell = ws.cell(row=row_num+1, column=1, value="TOTAL INGRESOS:")
        total_cell.font = Font(bold=True, color="00B050")
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.border = border
        
        ws.cell(row=row_num+1, column=7, value=total_ingresos).border = border
        ws.cell(row=row_num+1, column=7).font = Font(bold=True, color="00B050")
        ws.cell(row=row_num+1, column=7).number_format = '#,##0.00'
        
        ws.merge_cells(f'A{row_num+2}:F{row_num+2}')
        total_cell = ws.cell(row=row_num+2, column=1, value="TOTAL EGRESOS:")
        total_cell.font = Font(bold=True, color="FF0000")
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.border = border
        
        ws.cell(row=row_num+2, column=7, value=total_egresos).border = border
        ws.cell(row=row_num+2, column=7).font = Font(bold=True, color="FF0000")
        ws.cell(row=row_num+2, column=7).number_format = '#,##0.00'
        
        ws.merge_cells(f'A{row_num+3}:F{row_num+3}')
        balance = total_ingresos - total_egresos
        balance_color = "00B050" if balance >= 0 else "FF0000"
        total_cell = ws.cell(row=row_num+3, column=1, value="BALANCE FINAL:")
        total_cell.font = Font(bold=True, color=balance_color)
        total_cell.alignment = Alignment(horizontal='right')
        total_cell.border = border
        
        ws.cell(row=row_num+3, column=7, value=balance).border = border
        ws.cell(row=row_num+3, column=7).font = Font(bold=True, color=balance_color, size=14)
        ws.cell(row=row_num+3, column=7).number_format = '#,##0.00'
        
        # Ajustar anchos de columnas
        column_widths = [12, 10, 12, 35, 20, 12, 15, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_caja.xlsx"'
        
        wb.save(response)
        return response

    def _determinar_metodo_pago(self, pago):
        """Determina el método de pago basado en los montos"""
        if pago.yape > 0 and pago.efectivo > 0:
            return 'mixto'
        elif pago.yape > 0:
            return 'yape'
        else:
            return 'efectivo'
        
# In your views.py - add this to your existing ViewSet or create a new one
class POSViewSet(viewsets.ViewSet):
    
    @action(detail=False, methods=['post'])
    def registrar_salida_pos(self, request):
        try:
            from . import serializer as app_serializer
            
            pos_serializer = app_serializer.POSSalidaSerializer(data=request.data)
            if not pos_serializer.is_valid():
                return Response(pos_serializer.errors, status=400)
            
            data = pos_serializer.validated_data
            
            # Get the worker (should be Jefe de Planta or Asistente)
            trabajador = models.Trabajador.objects.get(id_trabajador=data['id_trabajador'])
            
            # Get the product
            producto = models.Producto.objects.get(id_producto=data['id_producto'])
            
            # Calculate values
            multiplicar_por = 2
            total_cancelar = data['cantidad'] * float(multiplicar_por)
            
            # 🔥 CORRECCIÓN: Usar el nombre correcto "Completada" y "Pagado"
            try:
                estado_salida = models.EstadoSalida.objects.get(nom_estado_salida="Completada")  # 🔥 Cambiado a "Completada"
            except models.EstadoSalida.DoesNotExist:
                # Si no existe, usar el que tenga ID 2
                estado_salida = models.EstadoSalida.objects.get(id_estado_salida=2)
            
            try:
                estado_pago = models.EstadoPago.objects.get(nom_estado="Pagado")
            except models.EstadoPago.DoesNotExist:
                # Si no existe, usar el primero disponible
                estado_pago = models.EstadoPago.objects.first()
            
            print(f"🔧 Usando EstadoSalida: {estado_salida.nom_estado_salida} (ID: {estado_salida.id_estado_salida})")
            print(f"🔧 Usando EstadoPago: {estado_pago.nom_estado} (ID: {estado_pago.id_estado})")
            
            # Create the Salida record
            salida = models.Salida.objects.create(
                id_trabajador=trabajador,
                id_producto=producto,
                cantidad=data['cantidad'],
                multiplicar_por=multiplicar_por,
                total_cancelar=total_cancelar,
                id_estado_salida=estado_salida,
                id_estado_pago=estado_pago,
                fecha=timezone.now().date(),
                hora=timezone.now().time()
            )
            
            # Create payment record
            pago = models.Pago.objects.create(
                efectivo=total_cancelar if data['metodo_pago'] in ['efectivo', 'mixto'] else 0,
                yape=total_cancelar if data['metodo_pago'] in ['yape', 'mixto'] else 0
            )
            
            # Get or create a default detail
            detalle_default = models.Detalle.objects.first()
            if not detalle_default:
                tipo_detalle_default = models.TipoDetalle.objects.first()
                detalle_default = models.Detalle.objects.create(
                    id_tipo_detalle=tipo_detalle_default,
                    descripcion="Detalle automático para POS"
                )
            
            # Create retorno record (since it's POS, it's immediately returned/paid)
            retorno = models.Retorno.objects.create(
                id_salida=salida,
                cantidad=data['cantidad'],
                id_pago=pago,
                total_cancelado=total_cancelar,
                id_detalle=detalle_default
            )
            
            # If client is provided, also create a Venta record
            if data.get('id_cliente'):
                venta = models.Venta.objects.create(
                    id_producto=producto,
                    cantidad=data['cantidad'],
                    precio_v=multiplicar_por,
                    cobra_de=multiplicar_por,
                    total_cancelar=total_cancelar,
                    id_cliente=models.Cliente.objects.get(id_cliente=data['id_cliente']),
                    fecha=timezone.now().date(),
                    hora=timezone.now().date(),
                    id_pago=pago,
                    id_estado=estado_pago,
                    total_cancelado=total_cancelar,
                    id_trabajador=trabajador
                )
            
            return Response({
                "message": "Salida POS registrada exitosamente",
                "salida_id": salida.id_salida,
                "producto": producto.nom_producto,
                "cantidad": data['cantidad'],
                "total": float(total_cancelar),
                "metodo_pago": data['metodo_pago']
            }, status=201)
            
        except models.Trabajador.DoesNotExist:
            return Response({"error": "Trabajador no encontrado"}, status=400)
        except models.Producto.DoesNotExist:
            return Response({"error": "Producto no encontrado"}, status=400)
        except Exception as e:
            print(f"Error registrando salida POS: {str(e)}")
            return Response({"error": str(e)}, status=500)
    
    # In your views.py - add these to existing ViewSets
    @action(detail=False, methods=['get'])
    def trabajadores_planta(self, request):
        """Get only plant workers (Jefe de Planta and Asistente)"""
        trabajadores = models.Trabajador.objects.filter(
            id_tipo_trabajador__in=[1, 2]  # IDs for Jefe de Planta and Asistente
        ).select_related('id_persona')
        
        datos = []
        for trab in trabajadores:
            datos.append({
                'id': trab.id_trabajador,
                'nombre': trab.id_persona.nombre_p,
                'apellido': trab.id_persona.apellido_p,
                'tipo_trabajador': trab.id_tipo_trabajador.nom_tt
            })
        
        return Response(datos)

    @action(detail=False, methods=['get'])
    def clientes_lista(self, request):
        """Get all clients"""
        clientes = models.Cliente.objects.all()
        
        datos = []
        for cliente in clientes:
            datos.append({
                'id_cliente': cliente.id_cliente,
                'nombre_cliente': cliente.nombre_cliente,
                'numero': cliente.numero,
                'direccion': cliente.direccion
            })
        
        return Response(datos)
    
# Add this to your views.py if ClienteViewSet doesn't exist
class ClienteViewSet(viewsets.ModelViewSet):
    queryset = models.Cliente.objects.all()
    
    @action(detail=False, methods=['get'])
    def clientes_lista(self, request):
        """Get all clients"""
        clientes = self.queryset.all()
        
        datos = []
        for cliente in clientes:
            datos.append({
                'id_cliente': cliente.id_cliente,
                'nombre_cliente': cliente.nombre_cliente,
                'numero': cliente.numero,
                'direccion': cliente.direccion
            })
        
        return Response(datos)
    

class SalidaVentaHibridaViewSet(viewsets.ViewSet):
    
    def list(self, request):
        """Obtener todas las operaciones: salidas + ventas POS"""
        # Obtener salidas
        salidas = models.Salida.objects.all().order_by('-fecha', '-hora')
        salidas_data = serializer.SalidaVentaHibridaSerializer(salidas, many=True).data
        
        # Obtener ventas POS (puedes filtrar por algún criterio si es necesario)
        ventas_pos = models.Venta.objects.all().order_by('-fecha', '-hora')
        ventas_data = serializer.VentaPOSSerializer(ventas_pos, many=True).data
        
        # Combinar y ordenar por fecha/hora
        todas_operaciones = salidas_data + ventas_data
        todas_operaciones.sort(key=lambda x: (x['fecha'], x['hora']), reverse=True)
        
        return Response(todas_operaciones)
    
    @action(detail=False, methods=['get'])
    def operaciones_hoy(self, request):
        """Obtener operaciones de hoy"""
        hoy = timezone.now().date()
        
        salidas_hoy = models.Salida.objects.filter(fecha=hoy)
        salidas_data = serializer.SalidaVentaHibridaSerializer(salidas_hoy, many=True).data
        
        ventas_hoy = models.Venta.objects.filter(fecha=hoy)
        ventas_data = serializer.VentaPOSSerializer(ventas_hoy, many=True).data
        
        todas_operaciones = salidas_data + ventas_data
        todas_operaciones.sort(key=lambda x: x['hora'], reverse=True)
        
        return Response(todas_operaciones)
    
# En tu views.py - REEMPLAZAR el ReportesViewSet actual

class ReportesViewSet(viewsets.ViewSet):

    @action(detail=False, methods=['post'])
    def generar_vista_previa(self, request):
        """Endpoint para generar vista previa del reporte"""
        try:
            print("🎯 INICIANDO VISTA PREVIA...")
            
            from . import serializer as app_serializer
            reporte_serializer = app_serializer.ReporteFiltrosSerializer(data=request.data)
            
            if not reporte_serializer.is_valid():
                print("❌ Error de validación:", reporte_serializer.errors)
                return Response(reporte_serializer.errors, status=400)
            
            data = reporte_serializer.validated_data
            tipo_reporte = data.get('tipo_reporte')
            
            print(f"📊 Tipo de reporte para vista previa: {tipo_reporte}")

            # Generar la vista previa según el tipo
            if tipo_reporte == 'ventas':
                return self._generar_vista_previa_ventas(data)
            elif tipo_reporte == 'entregas':
                return self._generar_vista_previa_entregas(data)
            elif tipo_reporte == 'trabajadores':
                return self._generar_vista_previa_trabajadores(data)
            elif tipo_reporte == 'productos':
                return self._generar_vista_previa_productos(data)
            else:  # completo
                return self._generar_vista_previa_completo(data)
            
        except Exception as e:
            print(f"❌ ERROR en vista previa: {str(e)}")
            import traceback
            print(f"❌ TRACEBACK: {traceback.format_exc()}")
            return Response(
                {"error": f"Error interno: {str(e)}"}, 
                status=500
            )

    def _generar_vista_previa_ventas(self, filtros):
        """Vista previa para reporte de ventas"""
        ventas = models.Venta.objects.all().select_related(
            'id_producto', 'id_cliente', 'id_trabajador', 'id_pago'
        )
        
        # Aplicar filtros
        if filtros.get('fecha_inicio'):
            ventas = ventas.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            ventas = ventas.filter(fecha__lte=filtros['fecha_fin'])
        if filtros.get('id_trabajador'):
            ventas = ventas.filter(id_trabajador=filtros['id_trabajador'])
        if filtros.get('id_producto'):
            ventas = ventas.filter(id_producto=filtros['id_producto'])
        
        # Datos para tabla
        datos_tabla = []
        for venta in ventas[:10]:  # Solo primeros 10 para vista previa
            metodo = 'efectivo'
            if venta.id_pago:
                if venta.id_pago.yape > 0 and venta.id_pago.efectivo > 0:
                    metodo = 'mixto'
                elif venta.id_pago.yape > 0:
                    metodo = 'yape'
            
            datos_tabla.append({
                'fecha': venta.fecha.strftime('%Y-%m-%d'),
                'hora': str(venta.hora),
                'producto': venta.id_producto.nom_producto,
                'cantidad': venta.cantidad,
                'precio_unitario': float(venta.precio_v),
                'total': float(venta.total_cancelado),
                'cliente': venta.id_cliente.nombre_cliente if venta.id_cliente else 'Cliente eventual',
                'vendedor': f"{venta.id_trabajador.id_persona.nombre_p} {venta.id_trabajador.id_persona.apellido_p}",
                'metodo_pago': metodo,
            })
        
        # Datos para gráfico (ventas por producto)
        ventas_por_producto = ventas.values('id_producto__nom_producto').annotate(
            total=Sum('total_cancelado')
        ).order_by('-total')[:5]
        
        grafico_ventas = [
            {
                'producto': item['id_producto__nom_producto'],
                'total': float(item['total'])
            }
            for item in ventas_por_producto
        ]
        
        return Response({
            'total_registros': ventas.count(),
            'total_ventas': float(ventas.aggregate(Sum('total_cancelado'))['total_cancelado__sum'] or 0),
            'headers': ['FECHA', 'HORA', 'PRODUCTO', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL', 'CLIENTE', 'VENDEDOR', 'MÉTODO PAGO'],
            'datos': datos_tabla,
            'grafico_ventas': grafico_ventas
        })

    def _generar_vista_previa_entregas(self, filtros):
        """Vista previa para reporte de entregas"""
        salidas = models.Salida.objects.all().select_related(
            'id_producto', 'id_trabajador', 'id_estado_salida', 'id_estado_pago'
        )
        
        # Aplicar filtros
        if filtros.get('fecha_inicio'):
            salidas = salidas.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            salidas = salidas.filter(fecha__lte=filtros['fecha_fin'])
        if filtros.get('id_trabajador'):
            salidas = salidas.filter(id_trabajador=filtros['id_trabajador'])
        if filtros.get('id_producto'):
            salidas = salidas.filter(id_producto=filtros['id_producto'])
        
        # Datos para tabla
        datos_tabla = []
        for salida in salidas[:10]:
            datos_tabla.append({
                'fecha': salida.fecha.strftime('%Y-%m-%d'),
                'hora': str(salida.hora),
                'producto': salida.id_producto.nom_producto,
                'cantidad': salida.cantidad,
                'precio_unitario': float(salida.multiplicar_por),
                'total_esperado': float(salida.total_cancelar),
                'repartidor': f"{salida.id_trabajador.id_persona.nombre_p} {salida.id_trabajador.id_persona.apellido_p}",
                'estado_entrega': salida.id_estado_salida.nom_estado_salida,
                'estado_pago': salida.id_estado_pago.nom_estado,
            })
        
        # Datos para gráfico (estados de entrega)
        estados_entrega = salidas.values('id_estado_salida__nom_estado_salida').annotate(
            cantidad=Count('id_salida')
        )
        
        grafico_entregas = [
            {
                'estado': item['id_estado_salida__nom_estado_salida'],
                'cantidad': item['cantidad']
            }
            for item in estados_entrega
        ]
        
        return Response({
            'total_registros': salidas.count(),
            'headers': ['FECHA', 'HORA', 'PRODUCTO', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL ESPERADO', 'REPARTIDOR', 'ESTADO ENTREGA', 'ESTADO PAGO'],
            'datos': datos_tabla,
            'grafico_entregas': grafico_entregas
        })

    def _generar_vista_previa_trabajadores(self, filtros):
        """Vista previa para reporte de trabajadores"""
        trabajadores = models.Trabajador.objects.all().select_related(
            'id_persona', 'id_tipo_trabajador'
        )
        
        if filtros.get('id_trabajador'):
            trabajadores = trabajadores.filter(id_trabajador=filtros['id_trabajador'])
        
        datos_tabla = []
        grafico_trabajadores = []
        
        for trabajador in trabajadores[:10]:
            ventas_trabajador = models.Venta.objects.filter(id_trabajador=trabajador)
            salidas_trabajador = models.Salida.objects.filter(id_trabajador=trabajador)
            
            if filtros.get('fecha_inicio'):
                ventas_trabajador = ventas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
                salidas_trabajador = salidas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
            if filtros.get('fecha_fin'):
                ventas_trabajador = ventas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
                salidas_trabajador = salidas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
            
            total_ventas = ventas_trabajador.aggregate(Sum('total_cancelado'))['total_cancelado__sum'] or 0
            total_entregas = salidas_trabajador.count()
            entregas_completadas = salidas_trabajador.filter(id_estado_salida=2).count()
            
            eficiencia = 0
            if total_entregas > 0:
                eficiencia = (entregas_completadas / total_entregas) * 100
            
            datos_tabla.append({
                'nombre': f"{trabajador.id_persona.nombre_p} {trabajador.id_persona.apellido_p}",
                'dni': trabajador.id_persona.dni_p,
                'tipo_trabajador': trabajador.id_tipo_trabajador.nom_tt,
                'total_ventas': float(total_ventas),
                'total_entregas': total_entregas,
                'entregas_completadas': entregas_completadas,
                'eficiencia': round(eficiencia, 2),
            })
            
            grafico_trabajadores.append({
                'nombre': f"{trabajador.id_persona.nombre_p} {trabajador.id_persona.apellido_p}",
                'eficiencia': round(eficiencia, 2)
            })
        
        return Response({
            'total_registros': trabajadores.count(),
            'promedio_eficiencia': round(sum([t['eficiencia'] for t in grafico_trabajadores]) / len(grafico_trabajadores) if grafico_trabajadores else 0, 2),
            'headers': ['NOMBRE', 'DNI', 'TIPO TRABAJADOR', 'TOTAL VENTAS', 'TOTAL ENTREGAS', 'ENTREGAS COMPLETADAS', 'EFICIENCIA (%)'],
            'datos': datos_tabla,
            'grafico_trabajadores': grafico_trabajadores
        })

    def _generar_vista_previa_productos(self, filtros):
        """Vista previa para reporte de productos"""
        from django.db.models import Count, Sum
        
        ventas_por_producto = models.Venta.objects.values(
            'id_producto', 
            'id_producto__nom_producto'
        ).annotate(
            total_vendido=Sum('cantidad'),
            total_ingresos=Sum('total_cancelado'),
            num_ventas=Count('id_venta')
        ).order_by('-total_vendido')
        
        if filtros.get('fecha_inicio'):
            ventas_por_producto = ventas_por_producto.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            ventas_por_producto = ventas_por_producto.filter(fecha__lte=filtros['fecha_fin'])
        if filtros.get('id_producto'):
            ventas_por_producto = ventas_por_producto.filter(id_producto=filtros['id_producto'])
        
        datos_tabla = []
        for item in ventas_por_producto[:10]:
            producto = models.Producto.objects.get(id_producto=item['id_producto'])
            precio_promedio = float(item['total_ingresos']) / item['total_vendido'] if item['total_vendido'] > 0 else 0
            
            datos_tabla.append({
                'producto': item['id_producto__nom_producto'],
                'total_vendido': item['total_vendido'],
                'total_ingresos': float(item['total_ingresos']),
                'num_ventas': item['num_ventas'],
                'tipo_producto': producto.id_tipo_producto.nom_tipo_p,
                'precio_promedio': round(precio_promedio, 2),
            })
        
        return Response({
            'total_registros': ventas_por_producto.count(),
            'headers': ['PRODUCTO', 'TOTAL VENDIDO', 'TOTAL INGRESOS', 'NÚMERO VENTAS', 'TIPO PRODUCTO', 'PRECIO PROMEDIO'],
            'datos': datos_tabla
        })

    def _generar_vista_previa_completo(self, filtros):
        """Vista previa para reporte completo (usa ventas como base)"""
        return self._generar_vista_previa_ventas(filtros)
    
    @action(detail=False, methods=['post'])
    def generar_reporte_flexible(self, request):
        try:
            print("🎯 INICIANDO GENERACIÓN DE REPORTE FLEXIBLE...")
            
            from . import serializer as app_serializer
            reporte_serializer = app_serializer.ReporteFiltrosSerializer(data=request.data)
            
            if not reporte_serializer.is_valid():
                print("❌ Error de validación:", reporte_serializer.errors)
                return Response(reporte_serializer.errors, status=400)
            
            data = reporte_serializer.validated_data
            tipo_reporte = data.get('tipo_reporte')
            fecha_inicio = data.get('fecha_inicio')
            fecha_fin = data.get('fecha_fin')
            id_trabajador = data.get('id_trabajador')
            id_producto = data.get('id_producto')
            metodo_pago = data.get('metodo_pago', 'todos')
            
            print(f"📊 Tipo de reporte: {tipo_reporte}")
            print(f"📅 Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")
            print(f"👤 Trabajador ID: {id_trabajador}")
            print(f"📦 Producto ID: {id_producto}")

            # Generar el reporte según el tipo
            if tipo_reporte == 'ventas':
                return self._generar_reporte_ventas(data)
            elif tipo_reporte == 'entregas':
                return self._generar_reporte_entregas(data)
            elif tipo_reporte == 'trabajadores':
                return self._generar_reporte_trabajadores(data)
            elif tipo_reporte == 'productos':
                return self._generar_reporte_productos(data)
            else:  # completo
                return self._generar_reporte_completo(data)
            
        except Exception as e:
            print(f"❌ ERROR: {str(e)}")
            import traceback
            print(f"❌ TRACEBACK: {traceback.format_exc()}")
            return Response(
                {"error": f"Error interno: {str(e)}"}, 
                status=500
            )

    def _generar_reporte_ventas(self, filtros):
        """Generar reporte de ventas con filtros flexibles"""
        # Aplicar filtros base
        ventas = models.Venta.objects.all().select_related(
            'id_producto', 'id_cliente', 'id_trabajador', 'id_pago'
        )
        
        # Aplicar filtros de fecha
        if filtros.get('fecha_inicio'):
            ventas = ventas.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            ventas = ventas.filter(fecha__lte=filtros['fecha_fin'])
            
        # Aplicar filtro de trabajador
        if filtros.get('id_trabajador'):
            ventas = ventas.filter(id_trabajador=filtros['id_trabajador'])
            
        # Aplicar filtro de producto
        if filtros.get('id_producto'):
            ventas = ventas.filter(id_producto=filtros['id_producto'])
        
        # Preparar datos para Excel
        datos_reporte = []
        for venta in ventas:
            # Determinar método de pago
            metodo = 'efectivo'
            if venta.id_pago:
                if venta.id_pago.yape > 0 and venta.id_pago.efectivo > 0:
                    metodo = 'mixto'
                elif venta.id_pago.yape > 0:
                    metodo = 'yape'
            
            datos_reporte.append({
                'fecha': venta.fecha,
                'hora': venta.hora,
                'producto': venta.id_producto.nom_producto,
                'cantidad': venta.cantidad,
                'precio_unitario': float(venta.precio_v),
                'total': float(venta.total_cancelado),
                'cliente': venta.id_cliente.nombre_cliente if venta.id_cliente else 'Cliente eventual',
                'vendedor': f"{venta.id_trabajador.id_persona.nombre_p} {venta.id_trabajador.id_persona.apellido_p}",
                'metodo_pago': metodo,
                'estado': venta.id_estado.nom_estado
            })
        
        return self._crear_excel_ventas(datos_reporte, filtros)

    def _generar_reporte_entregas(self, filtros):
        """Generar reporte de entregas con filtros flexibles"""
        # Aplicar filtros base
        salidas = models.Salida.objects.all().select_related(
            'id_producto', 'id_trabajador', 'id_estado_salida', 'id_estado_pago'
        )
        
        # Aplicar filtros de fecha
        if filtros.get('fecha_inicio'):
            salidas = salidas.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            salidas = salidas.filter(fecha__lte=filtros['fecha_fin'])
            
        # Aplicar filtro de trabajador
        if filtros.get('id_trabajador'):
            salidas = salidas.filter(id_trabajador=filtros['id_trabajador'])
            
        # Aplicar filtro de producto
        if filtros.get('id_producto'):
            salidas = salidas.filter(id_producto=filtros['id_producto'])
        
        # Preparar datos para Excel
        datos_reporte = []
        for salida in salidas:
            datos_reporte.append({
                'fecha': salida.fecha,
                'hora': salida.hora,
                'producto': salida.id_producto.nom_producto,
                'cantidad': salida.cantidad,
                'precio_unitario': float(salida.multiplicar_por),
                'total_esperado': float(salida.total_cancelar),
                'repartidor': f"{salida.id_trabajador.id_persona.nombre_p} {salida.id_trabajador.id_persona.apellido_p}",
                'estado_entrega': salida.id_estado_salida.nom_estado_salida,
                'estado_pago': salida.id_estado_pago.nom_estado,
                'tipo_trabajador': salida.id_trabajador.id_tipo_trabajador.nom_tt
            })
        
        return self._crear_excel_entregas(datos_reporte, filtros)

    def _generar_reporte_trabajadores(self, filtros):
        """Generar reporte de desempeño de trabajadores"""
        trabajadores = models.Trabajador.objects.all().select_related(
            'id_persona', 'id_tipo_trabajador'
        )
        
        # Aplicar filtro por tipo de trabajador si se especifica
        if filtros.get('id_trabajador'):
            trabajadores = trabajadores.filter(id_trabajador=filtros['id_trabajador'])
        
        datos_reporte = []
        for trabajador in trabajadores:
            # Estadísticas del trabajador
            ventas_trabajador = models.Venta.objects.filter(id_trabajador=trabajador)
            salidas_trabajador = models.Salida.objects.filter(id_trabajador=trabajador)
            
            # Aplicar filtros de fecha si existen
            if filtros.get('fecha_inicio'):
                ventas_trabajador = ventas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
                salidas_trabajador = salidas_trabajador.filter(fecha__gte=filtros['fecha_inicio'])
            if filtros.get('fecha_fin'):
                ventas_trabajador = ventas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
                salidas_trabajador = salidas_trabajador.filter(fecha__lte=filtros['fecha_fin'])
            
            total_ventas = ventas_trabajador.aggregate(Sum('total_cancelado'))['total_cancelado__sum'] or 0
            total_entregas = salidas_trabajador.count()
            entregas_completadas = salidas_trabajador.filter(id_estado_salida=2).count()
            
            eficiencia = 0
            if total_entregas > 0:
                eficiencia = (entregas_completadas / total_entregas) * 100
            
            datos_reporte.append({
                'nombre': f"{trabajador.id_persona.nombre_p} {trabajador.id_persona.apellido_p}",
                'dni': trabajador.id_persona.dni_p,
                'tipo_trabajador': trabajador.id_tipo_trabajador.nom_tt,
                'total_ventas': float(total_ventas),
                'total_entregas': total_entregas,
                'entregas_completadas': entregas_completadas,
                'eficiencia': round(eficiencia, 2),
                'sueldo': float(trabajador.sueldo)
            })
        
        return self._crear_excel_trabajadores(datos_reporte, filtros)

    def _generar_reporte_productos(self, filtros):
        """Generar reporte de productos más vendidos"""
        from django.db.models import Count, Sum
        
        # Agrupar ventas por producto
        ventas_por_producto = models.Venta.objects.values(
            'id_producto', 
            'id_producto__nom_producto'
        ).annotate(
            total_vendido=Sum('cantidad'),
            total_ingresos=Sum('total_cancelado'),
            num_ventas=Count('id_venta')
        ).order_by('-total_vendido')
        
        # Aplicar filtros de fecha
        if filtros.get('fecha_inicio'):
            ventas_por_producto = ventas_por_producto.filter(fecha__gte=filtros['fecha_inicio'])
        if filtros.get('fecha_fin'):
            ventas_por_producto = ventas_por_producto.filter(fecha__lte=filtros['fecha_fin'])
            
        # Aplicar filtro de producto específico
        if filtros.get('id_producto'):
            ventas_por_producto = ventas_por_producto.filter(id_producto=filtros['id_producto'])
        
        datos_reporte = []
        for item in ventas_por_producto:
            producto = models.Producto.objects.get(id_producto=item['id_producto'])
            datos_reporte.append({
                'producto': item['id_producto__nom_producto'],
                'total_vendido': item['total_vendido'],
                'total_ingresos': float(item['total_ingresos']),
                'num_ventas': item['num_ventas'],
                'tipo_producto': producto.id_tipo_producto.nom_tipo_p,
                'precio_promedio': float(item['total_ingresos']) / item['total_vendido'] if item['total_vendido'] > 0 else 0
            })
        
        return self._crear_excel_productos(datos_reporte, filtros)

    def _generar_reporte_completo(self, filtros):
        """Generar reporte completo combinado"""
        # Aquí puedes combinar datos de múltiples reportes
        # Por simplicidad, usaremos el de ventas como base
        return self._generar_reporte_ventas(filtros)

    # Métodos para crear Excel (similar al de caja pero adaptado)
    def _crear_excel_ventas(self, datos, filtros):
        return self._crear_excel_profesional(
            datos, 
            "REPORTE DE VENTAS - YACU SELVA", 
            filtros,
            ['FECHA', 'HORA', 'PRODUCTO', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL', 'CLIENTE', 'VENDEDOR', 'MÉTODO PAGO', 'ESTADO']
        )

    def _crear_excel_entregas(self, datos, filtros):
        return self._crear_excel_profesional(
            datos,
            "REPORTE DE ENTREGAS - YACU SELVA",
            filtros,
            ['FECHA', 'HORA', 'PRODUCTO', 'CANTIDAD', 'PRECIO UNIT.', 'TOTAL ESPERADO', 'REPARTIDOR', 'ESTADO ENTREGA', 'ESTADO PAGO', 'TIPO TRABAJADOR']
        )

    def _crear_excel_trabajadores(self, datos, filtros):
        return self._crear_excel_profesional(
            datos,
            "REPORTE DE TRABAJADORES - YACU SELVA",
            filtros,
            ['NOMBRE', 'DNI', 'TIPO TRABAJADOR', 'TOTAL VENTAS', 'TOTAL ENTREGAS', 'ENTREGAS COMPLETADAS', 'EFICIENCIA (%)', 'SUELDO']
        )

    def _crear_excel_productos(self, datos, filtros):
        return self._crear_excel_profesional(
            datos,
            "REPORTE DE PRODUCTOS - YACU SELVA",
            filtros,
            ['PRODUCTO', 'TOTAL VENDIDO', 'TOTAL INGRESOS', 'NÚMERO VENTAS', 'TIPO PRODUCTO', 'PRECIO PROMEDIO']
        )

    def _crear_excel_profesional(self, datos, titulo, filtros, headers):
        """Método base para crear Excel profesional (similar al de caja)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos (usar los mismos que en el reporte de caja)
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2C5F2D", end_color="2C5F2D", fill_type="solid")
        title_font = Font(bold=True, size=16, color="2C5F2D")
        subheader_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título principal
        ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
        ws['A1'] = titulo
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Información del reporte
        ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
        periodo = "PERIODO: "
        if filtros.get('fecha_inicio') and filtros.get('fecha_fin'):
            periodo += f"{filtros['fecha_inicio']} al {filtros['fecha_fin']}"
        elif filtros.get('fecha_inicio'):
            periodo += f"Desde {filtros['fecha_inicio']}"
        elif filtros.get('fecha_fin'):
            periodo += f"Hasta {filtros['fecha_fin']}"
        else:
            periodo += "Todos los registros"
            
        if filtros.get('id_trabajador'):
            periodo += f" | TRABAJADOR: {filtros['id_trabajador']}"
        if filtros.get('id_producto'):
            periodo += f" | PRODUCTO: {filtros['id_producto']}"
            
        ws['A2'] = periodo
        ws['A2'].font = subheader_font
        ws['A2'].alignment = center_align
        
        # Fecha de generación
        ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
        ws['A3'] = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].alignment = center_align
        
        # Espacio
        ws.row_dimensions[4].height = 5
        
        # Encabezados de columnas
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Datos
        row_num = 6
        for dato in datos:
            for col, (key, value) in enumerate(dato.items(), 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                # Formatear números
                if isinstance(value, (int, float)) and col > 1:
                    cell.number_format = '#,##0.00'
            row_num += 1
        
        # Ajustar anchos de columnas
        for i in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 15
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Nombre del archivo
        fecha = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f"reporte_{filtros['tipo_reporte']}_{fecha}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        
        wb.save(response)
        return response